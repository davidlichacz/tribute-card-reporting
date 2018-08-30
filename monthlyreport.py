import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from functions import adjust_width
import os, subprocess, sys

def summary_stats(dataframe, pulled, sent, month, year, path):
    # Use openpyxl package for inputing summary statistics.
    wb_cards = openpyxl.Workbook()
    ws = wb_cards.active

    for row in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(row)

    # Enter summary statistics and labels.
    last_row = ws.max_row
    mean_pulled = round(np.mean(pulled), 2)
    mean_sent = round(np.mean(sent), 2)
    max_pulled = max(pulled)
    max_sent = max(sent)

    ws['I'+str(last_row+2)] = 'Average:'
    ws['I'+str(last_row+2)].alignment = openpyxl.styles.Alignment(horizontal='right')

    ws['J'+str(last_row+2)] = mean_pulled
    ws['K'+str(last_row+2)] = mean_sent

    ws['I'+str(last_row+3)] = 'Maximum:'
    ws['I'+str(last_row+3)].alignment = openpyxl.styles.Alignment(horizontal='right')

    ws['J'+str(last_row+3)] = max_pulled
    ws['K'+str(last_row+3)] = max_sent

    adjust_width(ws)

    # Prepare, save and open final file.
    reportfile = f'{path}{month} {year}/{month} {year} Tribute Cards.xlsx'

    wb_cards.save(reportfile)

    return (reportfile, [last_row-1, mean_pulled, max_pulled, mean_sent, max_sent])

def open_file(file):
    if sys.platform.startswith('darwin'):
        subprocess.call(('open', file))
    elif os.name == 'nt':
        os.startfile(file)
    elif os.name == 'posix':
        subprocess.call(('xdg-open', file))