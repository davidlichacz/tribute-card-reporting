import glob
import pandas as pd
import bizdays
import os
import sys
import openpyxl
from numpy import mean
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess, os
import functions
import constants as c
from openpyxl.styles import Alignment

# Get month and year of report from user.  Ultimately this will be a selection from a drop-down list
# so error handling is omitted for now.
month = input('Enter month: ')
year = input('Enter year: ')

# Build list of files for the chosen month.
path = '/Users/davidlichacz/Tribute Spreadsheets/'
filepath = f'{path}{month} {year}/*.XLS'
filelist = glob.glob(filepath)

# If an error log from a previous running of the report exisits, delete it.
errorlog = f'{path}/Error Logs/{month} {year} error log.txt'
try:
    os.remove(errorlog)
except:
    pass

# Check if folder is empty. If it is, there is no reason to proceed.
if len(filelist) == 0:
    print(f'There are currently no cards processed for {month} {year}.')
    sys.exit()

# Initialize a calendar that will calculate the number of business days between two dates.
cal = bizdays.Calendar(c.holidays, ['Sunday', 'Saturday'])

# Create empty dataframe that will contain all card data.
cards = pd.DataFrame()

# Read each Excel file into the cards dataframe.
for filename in filelist:
    data = pd.read_excel(filename)
    error = False 
    # Check to see if spreadsheet has the correct structure for reporting.
    # If it does not, add an entry to the error log for further investigation.
    if (len(data.columns.values) != len(c.sheet_columns)) or not np.array_equal(data.columns.values[0:7], np.array(c.sheet_columns[0:7])):
        error = True
        log = functions.open_error_log(errorlog)
        log.write(f'{filename} is incompatible with tribute spreadsheets.\n')
    else:
        # Removes potential inconsistencies in manually entered column names.
        data.columns.values[7:] = c.sheet_columns[7:]
        # Isolate dates columns as they are the only ones important for calculations.
        dates = data[['Gift Date Added', 'Date Pulled', 'Date Sent']]
        # See if any of the dates are missing.
        nans = pd.isnull(dates).any(1)
        nans_rows = nans.index[nans == True]
        if len(nans_rows) != 0:
            error = True
            log = functions.open_error_log(errorlog)
            for row in nans_rows:
                log.write(f'Row {row+2} in {filename} is missing one or more dates.\n')
    if error == False:
        # Removes potential inconsistencies in manually entered column names.       
        cards = pd.concat([cards, data], ignore_index=True)
# If there were any errors found, end process so user can investigate and correct issues.
try:
    if log.mode == 'a+':
        print('There was a problem generating the report.  Please check error log.')
        log.close
        sys.exit()
except NameError:
    pass

# Convert date columns to strings. Necessary to calculate differences in business days.  
cards['Gift Date Added'] = cards['Gift Date Added'].astype(str)
cards['Date Pulled'] = cards['Date Pulled'].astype(str)
cards['Date Sent'] = cards['Date Sent'].astype(str)

# Convert Constituent ID column to string to improve readability of final spreadsheet.
cards['Constituent ID'] = cards['Constituent ID'].astype(str)

# Calculate differences in business days and insert them into dataframe.
rows = cards.shape[0]
pulled = []
sent = []

for k in (range(rows)):
    pulled.append(functions.bizdays_neg(cal, cards['Gift Date Added'][k], cards['Date Pulled'][k]))
    sent.append(functions.bizdays_neg(cal, cards['Gift Date Added'][k], cards['Date Sent'][k]))

cards.insert(loc=9, column='Business Days Until Pulled', value=pulled)
cards.insert(loc=10, column='Business Days Until Sent', value=sent)

# Sort dataframe for logical readability.
cards = cards.sort_values(by=['Gift Date Added', 'Tribute Card Type'])

# Use openpyxl package for inputing summary statistics.
wb_cards = openpyxl.Workbook()
ws = wb_cards.active

for row in dataframe_to_rows(cards, index=False, header=True):
    ws.append(row)


# Enter summary statistics and labels.
last_row = ws.max_row

ws['I'+str(last_row+2)] = 'Average:'
ws['I'+str(last_row+2)].alignment = Alignment(horizontal='right')

ws['J'+str(last_row+2)] = round(mean(pulled), 2)
ws['K'+str(last_row+2)] = round(mean(sent), 2)

ws['I'+str(last_row+3)] = 'Maximum:'
ws['I'+str(last_row+3)].alignment = Alignment(horizontal='right')

ws['J'+str(last_row+3)] = max(pulled)
ws['K'+str(last_row+3)] = max(sent)

functions.adjust_width(ws)

# Prepare, save and open final file.
reportfile = f'{path}{month} {year}/{month} {year} Tribute Cards.xlsx'

wb_cards.save(reportfile)

if sys.platform.startswith('darwin'):
    subprocess.call(('open', reportfile))
elif os.name == 'nt':
    os.startfile(reportfile)
elif os.name == 'posix':
    subprocess.call(('xdg-open', reportfile))