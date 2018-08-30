from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from functions import adjust_width
import matplotlib.pyplot as plt
from math import pi
from matplotlib.colors import Normalize
import matplotlib.cm as cm
from openpyxl.drawing.image import Image
import constants as c
import openpyxl


# List of columns for which data will be populated.
columns = ['B', 'C', 'D', 'E', 'F']


def build_sheet(ws, stats):
	# Function that creates a new Excel spreadsheet for the annual report when the April report is run.

	# List of column labels.
	column_labels = ['Total Cards Sent', 'Average Days Until Pulled', 'Maximum Days Until Pulled', 
					 'Average Days Until Sent', 'Maximum Days Until Sent']

	# Populate the first column with month names.
	for row in range(2, 14):
		cell = ws['A' + str(row)]
		cell.value = c.months[row-2]
		cell.alignment = Alignment(horizontal='right')

	# Add a totals row to row 14.
	ws['A14'] = 'Total:'
	ws['A14'].alignment = Alignment(horizontal='right')


	# Populate worksheet with values and format cells accordingly.
	for x in range(0, 5):
		# Column labels
		label = columns[x] + str(1)
		ws[label] = column_labels[x]
		ws[label].alignment = Alignment(horizontal='center')
		ws[label].font = Font(bold=True)

		# Data for April
		cell = columns[x] + str(2)
		ws[cell] = stats[x]
		ws[cell].alignment = Alignment(horizontal='center')
		if columns[x] == 'C' or columns[x] == 'E':
			ws[cell].number_format = '0.00'

		# Totals row.  Will be the same as April row since we only have one month's worth of data.
		total = columns[x] + str(14)
		ws[total].value = ws[cell].value
		ws[total].alignment = Alignment(horizontal='center')
		if columns[x] == 'C' or columns[x] == 'E':
			ws[total].number_format = '0.00'

	# Make spreadsheet more readable when opened.
	adjust_width(ws)

	# Return a tuple with a list of total cards, average pulled, and average sent as well as scalar values for average pulled
	# and average sent.  These values are necessary for creation of graph.
	return ([stats[0]], [stats[1]], [stats[3]], stats[1], stats[3])


def update_sheet(ws, month, stats):
	# Function that updates the annual spreadsheet when a report for any month after April is run.

	# Find the appropriate row in the spreadsheet for the given month.
	row = c.months.index(month) + 2
	# Create empty lists that will hold the data that will calculate annual averages.
	total_list, mean_pulled_list, max_pulled_list, mean_sent_list, max_sent_list = ([] for i in range(5))

	# Build lists from previous monthly data.
	for x in range(2, row):
		total_list.append(ws['B'+str(x)].value)
		mean_pulled_list.append(ws['C'+str(x)].value)
		max_pulled_list.append(ws['D'+str(x)].value)
		mean_sent_list.append(ws['E'+str(x)].value)
		max_sent_list.append(ws['F'+str(x)].value)


	# Add current month's data to lists.
	total_list.append(stats[0])
	mean_pulled_list.append(stats[1])
	max_pulled_list.append(stats[2])
	mean_sent_list.append(stats[3])
	max_sent_list.append(stats[4])


	# Calculate total number of cards sent for the year
	annual_total = sum(total_list)

	# Calculate annual averages for pulled and sent times.
	annual_mean_pulled = sum([x*y/annual_total for (x, y) in list(zip(total_list, mean_pulled_list))])
	annual_mean_sent = sum([x*y/annual_total for (x, y) in list(zip(total_list, mean_sent_list))])

	# Add annual totals to spreadsheet
	ws['B14'] = sum(total_list)
	ws['C14'] = annual_mean_pulled
	ws['D14'] = max(max_pulled_list)
	ws['E14'] = annual_mean_sent
	ws['F14'] = max(max_sent_list)

	# Update spreadsheet with new data.
	for x in range(0, 5):
		# Add current month's data to appropriate row.
		cell = columns[x] + str(row)
		ws[cell] = stats[x]
		ws[cell].alignment = Alignment(horizontal='center')
		if columns[x] == 'C' or columns[x] == 'E':
			ws[cell].number_format = '0.00'

		# Update annual total row.
		total = columns[x] + str(14)
		ws[total].alignment = Alignment(horizontal='center')
		if columns[x] == 'C' or columns[x] == 'E':
			ws[total].number_format = '0.00'

	# Return a tuple with a list of total cards, average pulled and average sent monthly,  as well as scalar values for average pulled annual
	# and average sent annual.  These values are necessary for creation of graph.
	return (total_list, mean_pulled_list, mean_sent_list, annual_mean_pulled, annual_mean_sent)


def create_graph(path, month, fiscal, graph_stats):
	# Function that creates a graph based on to-date data for tribute card fulfillment.

	# Calculate the number of months remaining in the year and create zero values for them.
	months_remaining = 12 - len(graph_stats[0])
	trailing_zeros = months_remaining*[0]

	# Create 12 sequential x-values.
	x_ticks = list(range(1, 13))

	# y-values are the list of monthly average sent values.
	y_sent = graph_stats[2]
	y_sent.extend(trailing_zeros)

	# The number of cards sent in a month will determine the size of the dot on the graph.
	num_cards = graph_stats[0]
	num_cards.extend(trailing_zeros)
	size = [pi*(x**2)/4000 for x in num_cards]

	# The color of each dot will be determined by the monthly averaged pulled values.
	# The data is normalized by a factor of 7 as 7 days is the maximum allowed for card fulfillment.
	# This way any average date pulled with a value of 7 or higher will be assigned the highest colour.
	pulled = graph_stats[1]
	pulled.extend(trailing_zeros)
	pulled_norm = [x/7 for x in pulled]

	# Initialize colour map and assign a color for the normalized pulled values.
	cmap = cm.get_cmap('bwr')
	colors = cmap(pulled_norm)

	# Necessary for normalizing legend for colour map.
	m = cm.ScalarMappable(cmap=cmap)
	m.set_array(range(0, 8))

	# Define variables for annual averages.
	pulled_avg = round(graph_stats[3], 2)
	sent_avg = round(graph_stats[4], 2)


	# Define text for annotation for graph, depending on if year is complete.
	text = 'Annual averages'
	if months_remaining > 0:
		text += ' to date'

	# Graph annotation
	textstr = '\n'.join([text, f'Pulled: {pulled_avg:.2f} days', f'Sent: {sent_avg:.2f} days'])


	# Graph details
	plt.scatter(x_ticks, y_sent, size, color=[cmap(color) for color in pulled_norm], edgecolors='black')
	plt.title(f'Tribute Card Fulfillment: {fiscal}')
	plt.xlabel('Month')
	plt.ylabel('Average Business Days Until Sent')
	plt.xticks(x_ticks, c.months_abbrev)
	# Horizontal line at y=7 since 7 is the target for card fulfillmeat.
	plt.axhline(7)
	plt.colorbar(m, label='Average Business Days Until Pulled').ax.set_yticklabels([0, 1, 2, 3, 4, 5, 6, '7+'])
	plt.figtext(1.12, 0.5, textstr, horizontalalignment='center', bbox=dict(boxstyle="round", facecolor='#D8D8D8',
                      ec="0.5", pad=0.5, alpha=1), fontweight='bold')

	plt.savefig(f'{path}Tribute Card Fulfillment {fiscal}.png', bbox_inches='tight')
	plt.clf()

def annual_report(month, fiscal, path, stats):
	# Function that complies the annual report.

	file = f'{path}Tribute Cards {fiscal}.xlsx'

	# If month is April, start a new annual report.
	if month == 'April':
		wb = openpyxl.Workbook()
		ws = wb.active
		graph_stats = build_sheet(ws, stats)
	# Otherwise, open file for the appropriate fiscal year.
	else:
		wb = openpyxl.load_workbook(filename=file)
		ws = wb.active
		graph_stats = update_sheet(ws, month, stats)


	create_graph(path, month, fiscal, graph_stats)

	# Add graph to spreadsheet.
	img = Image(f'{path}Tribute Card Fulfillment: {fiscal}.png')
	ws.add_image(img, 'B18')
	wb.save(file)
